#Import Packages
from tkinter import *
from tkinter.scrolledtext import ScrolledText
import random
import time
import ctypes
from copy import deepcopy

#Import created modules
from my_func import *


#**********************************************
#********************CLASSES*******************

class Team():
    def __init__(self,name,group):
        self.name = name
        self.group = group
        self.teamsplayed = []
        self.lastdate = None
        self.homeground = None
        
    def lastplayed(self,date):
        diff = subtractDates(date,self.lastdate)
        return diff
        

class Match():
    def __init__(self,team1,team2,group,date,time=None,venue=None):
        self.team1 = team1
        self.team2 = team2
        self.group = group
        self.date = date
        self.time = time
        self.venue = venue

    def getTeamNames(self):
        return [self.team1.name,self.team2.name]


class Schedule():
    def __init__(self,all_teams,all_matches,include_venues,include_homegrounds,include_dates,parallel_groups,include_timeslots):
        self.all_teams = all_teams
        self.all_matches = all_matches
        self.include_venues = include_venues
        self.include_homegrounds = include_homegrounds
        self.include_dates = include_dates
        self.parallel_groups = parallel_groups
        self.include_timeslots = include_timeslots
        self.groupnames = []
        for team in self.all_teams:
            if not (team.group in self.groupnames):
                self.groupnames.append(team.group)
    
    def get_grouplen(self):
        grouplen = 0
        for team in all_teams:
            if team.group == all_teams[0].group:
                grouplen+=1
        return grouplen

    def get_numOfGroups(self):
        return int(len(all_teams)/self.get_grouplen())

    def print_allmatches(self):
        #display all matches
        for m_index,m in enumerate(self.all_matches):
            m_date = m.date.get_date()
            if include_timeslots:
                m_time = m.time.getTimeString()
            else:
                m_time = None
            m_num  = m_index + 1
            m_group = m.group
            m_t1   = m.team1.name
            m_t2   = m.team2.name
            m_venue= m.venue
            print(m_date,m_time,'  MATCH',m_num,'GROUP:',m_group,' ',m_t1,'vs',m_t2,'at',m_venue)

    def print_teammatches(self):
        #display matches per team
        for team in self.all_teams:
            print('----------------------')
            print(team.name)
            print('----------------------')
            for m_index,m in enumerate(self.all_matches):
                m_date = m.date.get_date()
                if self.include_timeslots:
                    m_time = m.time.getTimeString()
                else:
                    m_time = None
                m_num  = m_index + 1
                m_t1   = m.team1.name
                m_t2   = m.team2.name
                m_venue= m.venue
                if team.name in m.getTeamNames():
                    home = 'Home' if team.homeground == m_venue else 'Away'
                    print(m_date,m_time,'  MATCH',m_num,m_t1,'vs',m_t2,'at',m_venue,home)

    def xl_output(self,name):
        #Import libraries
        import openpyxl
        from openpyxl import workbook, load_workbook, styles
        from openpyxl.utils import get_column_letter
        from datetime import datetime

        #--------Function to write a row--------
        def write_row():
            #Write data in row
            for c,column in enumerate(columns):
                col = c+1
                value = 'n/a'
                sheet[getcell(1,col)].alignment = styles.Alignment(horizontal='center')
                sheet[getcell(1,col)].value = columns[col-1]
                sheet[getcell(1,col)].font = styles.Font(bold=True)
                if column == 'Match Number':
                    value = m_num
                    sheet.column_dimensions[get_column_letter(col)].width = 15
                elif column == 'Date':
                    value = m_date
                    sheet.column_dimensions[get_column_letter(col)].width = 15
                elif column == 'Day':
                    value = m_day
                    sheet.column_dimensions[get_column_letter(col)].width = 15
                elif column == 'Time':
                    value = m_time
                elif column == 'Group':
                    value = m_group
                elif column == 'Team1':
                    value = m_t1
                    sheet.column_dimensions[get_column_letter(col)].width = 15
                elif column == 'Team2':
                    value = m_t2
                    sheet.column_dimensions[get_column_letter(col)].width = 15
                elif column == 'Venue':
                    value = m_venue
                    sheet.column_dimensions[get_column_letter(col)].width = 15
                
                try:
                    cell = sheet[getcell(row,col)]
                    cell.value = value
                    cell.alignment = styles.Alignment(horizontal='center')
                    if column=='Date':
                        cell.number_format = 'YYYY MMM DD'
                except:
                    print('Error when setting value for cell',str(row)+chr(col+64))
                    sheet[getcell(row,col)].value = str(value)
        #------------------------------------
        #Setup workbook
        filename = str(name)+'.xlsx'
        wb = openpyxl.Workbook()
        #------------------------------------
        #Sheet: All Matches
        sheet = wb.active
        sheet.title = 'All Matches'

        #Make a list of required columns
        columns = ['Match Number','Date','Day','Time','Group','Team1','Team2','Venue']
        if not self.include_dates:
            columns.remove('Date')
            columns.remove('Day')
        if not self.include_timeslots:
            columns.remove('Time')
        if self.get_numOfGroups() < 2:
            columns.remove('Group')
        if not self.include_venues:
            columns.remove('Venue')

        #Write rows
        for m_index,m in enumerate(self.all_matches):
            #Get data of current row
            year,month,day = m.date.get_date()
            year,month,day = str(year),str(month),str(day)
            m_date = datetime.strptime(year+'-'+month+'-'+day,'%Y-%m-%d')
            m_day = get_dayOfWeek(m.date)
            m_day = m_day[0].upper() + m_day[1:]
            if include_timeslots:
                m_time = m.time.getTimeString()
            else:
                m_time = None
            m_num  = m_index + 1
            m_group = m.group
            m_t1   = m.team1.name
            m_t2   = m.team2.name
            m_venue= m.venue
            row = m_index+2
            write_row()
            
        wb.save(filename)
        #------------------------------------
        #Create sheets for each groups
        for group in self.groupnames:
            wb.create_sheet('Group '+group)
            sheet = wb['Group '+group]

            #Make a list of required columns
            columns = ['Match Number','Date','Day','Time','Team1','Team2','Venue']
            if not self.include_dates:
                columns.remove('Date')
                columns.remove('Day')
            if not self.include_timeslots:
                columns.remove('Time')
            if not self.include_venues:
                columns.remove('Venue')

            row = 1
            #Write rows
            for m_index,m in enumerate(self.all_matches):
                if m.group != group:
                    continue
                row+=1
                #Get data of current row
                year,month,day = m.date.get_date()
                year,month,day = str(year),str(month),str(day)
                m_date = datetime.strptime(year+'-'+month+'-'+day,'%Y-%m-%d')
                m_day = get_dayOfWeek(m.date)
                m_day = m_day[0].upper() + m_day[1:]
                if include_timeslots:
                    m_time = m.time.getTimeString()
                else:
                    m_time = None
                m_num  = m_index + 1
                m_group = m.group
                m_t1   = m.team1.name
                m_t2   = m.team2.name
                m_venue= m.venue  
                write_row()
        wb.save(filename)
        #------------------------------------
        for team in self.all_teams:
            #Setup workbook
            filename = str(name+'_'+team.name)+'.xlsx'
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = team.name

            #Make a list of required columns
            columns = ['Match Number','Date','Day','Time','Team1','Team2','Venue']
            if not self.include_dates:
                columns.remove('Date')
                columns.remove('Day')
            if not self.include_timeslots:
                columns.remove('Time')
            if not self.include_venues:
                columns.remove('Venue')
            row = 1
            #Write rows
            for m_index,m in enumerate(self.all_matches):
                if m.team1.name != team.name and m.team2.name != team.name:
                    continue
                row+=1
                #Get data of current row
                year,month,day = m.date.get_date()
                year,month,day = str(year),str(month),str(day)
                m_date = datetime.strptime(year+'-'+month+'-'+day,'%Y-%m-%d')
                m_day = get_dayOfWeek(m.date)
                m_day = m_day[0].upper() + m_day[1:]
                if include_timeslots:
                    m_time = m.time.getTimeString()
                else:
                    m_time = None
                m_num  = m_index + 1
                m_group = m.group
                m_t1   = m.team1.name
                m_t2   = m.team2.name
                m_venue= m.venue  
                write_row()
            wb.save(filename)
        
        
        
        
        
                
        
        
    

#******************************************************
#********************SHORT FUNCTIONS*******************

def widgetColor(widget):
    widget_colors = {
'w':'Grey',
'frame_mainInput':'Grey50',
'frame_1':'grey85',
'frame_2':'grey85',
'frame_3':'grey85',
'frame_4':'grey85',
'frame_5':'grey85',
'btn_left':'grey85',
'btn_schedule':'grey85',
'status':'Grey50'
        }

    #return 'grey85'
    return widget_colors[widget]
    

def getcell(row,col):
    from openpyxl.utils import get_column_letter
    col_char = get_column_letter(col)
    return col_char + str(row)


#Takes a list of Team objects and returns a list of team names
#Useful for printing a list of teams as they are objects of class Team
def teamList(objectList):
    new = []
    if objectList is None:
        return None
    for team in objectList:
        name = team.name
        new.append(name)
    return new


#Takes a list of Team objects and sorts them by teamsplayed
def sortTeams(l):
    for p in range(len(l)-1):
        for i in range(0,len(l)-1-p):
            if len(l[i].teamsplayed) > len(l[i+1].teamsplayed):
                l[i],l[i+1] = l[i+1],l[i]


#Takes number of teams and finds expected total round robin matches
def expectedLen(teamlen):
    return int(teamlen*(teamlen-1)*(1/2))
    

#Resets teamsplayed for all teams in a list
def resetTeams(teams):
    for team in teams:
        team.teamsplayed = []


def sortMatches(match_list,sortby='date',reverse=False):
    r = reverse
    if sortby == 'date':
        match_list.sort(key=lambda m: m.date.get_date(), reverse=r )
    elif sortby == 'group':
        match_list.sort(key=lambda m: m.group, reverse=r )
    elif sortby == 'team':
        match_list.sort(key=lambda m: m.team1.name, reverse=r )

        
def set_teamlastdate():
    #'last played date' for each team is set to minimum value
    for group in groups:
        for team in group:
            team.lastdate = Date([1,1,1])


def set_homegrounds():
    for i,team in enumerate(all_teams):
        team.homeground = venues[i]


#----------GUI Functions----------
def filter_int(entry_variable):
    val = entry_variable.get()
    #Below block removes any non int character from the string
    if len(val)>1:
        for i in range(len(val)-2,-1,-1):
            try:
                int(val[i])
            except:
                val = val[:i]+val[i+1:]
        entry_variable.set(val)
            
    try:
        int(val[-1])
    except:
        #print('val:',val)
        entry_variable.set(val[:-1])
        #print('var_e_11 now:',var_e_11.get())


def entry_limiter(entry_variable,limit):
    #Limit number of characters in an Entry widget
    val = entry_variable.get()
    while len(val)>limit:
        val = val[:-1]
    entry_variable.set(val)


def popup_error(title,msg):
    ctypes.windll.user32.MessageBoxW(0,msg,title,0)






























#**********************************************************
#********************PROCEDURAL FUNCTIONS*******************
    
def check_numOfTeams():
    global num_of_teams
    global num_of_groups
    global teamnames
    global error_status
    
    num_of_teams = 0
    num_of_groups = 1

    #----------Get num_of_teams----------
    try:
        #Input raw text of all team names
        text = e_teams.get('1.0',END)
        teamnames = [line for line in text.split('\n')]
        #Strip off blank space from each line
        for i in range(len(teamnames)):
            teamnames[i] = teamnames[i].strip()
        #Remove empty lines
        for i,name in enumerate(teamnames.copy()):
            if name == '':
                teamnames.remove(name)
        
        num_of_teams = len(teamnames)
        if num_of_teams > 199:
            error_status = '"number of teams" should be in the range 2-199'
        elif num_of_teams < 2:
            error_status = 'Enter one team name per line'

        lbl_numOfTeamsAndGroups.configure(text='Number of teams: '+str(num_of_teams)+'\nNumber of groups: '+str(num_of_groups))
    except Exception as e:
        print(e)


def check_numOfGroups():
    global num_of_teams
    global num_of_groups
    global error_status

    #----------Check num_of_groups----------
    empty_input = False
    invalid_input = True
    while 'Checking number of groups':
        #Input num_of_groups
        value = var_e_11.get()
        value = value.strip()
        if value == '':
            invalid_input=False
            empty_input=True
            break
        try:
            value = int(value)
            
        except:
            validation_msg = '"number of groups" should be an integer'
            break
        if value < 1 or ((value >= num_of_teams) and num_of_teams > 0):
            validation_msg = '"number of groups" should be less than "number of teams"'
            break
        elif value > 26:
            validation_msg = '"number of groups" too high! Enter value in the range 1-26'
            break

        if num_of_teams % value != 0:
            validation_msg = "Warning!\nnum. of teams '"+str(num_of_teams)+"' is not divisible by num. of groups '"+str(value)+"'"
            break
        else:
            num_of_groups = value
            invalid_input = False
            break
    
    if empty_input:
        error_status='"number of groups" cannot be empty'
    elif invalid_input:
        error_status=validation_msg
    else:
        error_status=''
        
    lbl_numOfTeamsAndGroups.configure(text='Number of teams: '+str(num_of_teams)+'\nNumber of groups: '+str(num_of_groups))


def process_teams():
    global grouplen
    global groups
    global all_teams
    global error_status

    #No two teams should have same names
    for i in range(len(teamnames)-1):
        for j in range(i+1,len(teamnames)):
            if teamnames[i] == teamnames[j]:
                error_status = 'The team name "'+str(teamnames[i])+'" is repeated!\nTwo team names cannot be same.'
                return

    #Find length of group (number of teams in one group)
    grouplen = int(num_of_teams / num_of_groups)

    #Create groups of teams
    groups = []
    all_teams = []
    for g in range(1,num_of_groups+1):
        current_group = []
        for n in range(1,grouplen+1):
            team = Team(teamnames[grouplen*(g-1)+n-1],chr(g+64))
            current_group.append(team)
            all_teams.append(team)
        groups.append(current_group)
    

def check_venues():
    global error_status
    global venues
    global include_homegrounds

    if not include_venues:
        return

    if var_21.get():
        include_homegrounds = True
    else:
        include_homegrounds = False
        
    num_of_venues = 0
    venues = []

    #----------Get num_of_venues----------
    try:
        #Input raw text of all venue names
        text = e_venues.get('1.0',END)
        venues = [line for line in text.split('\n')]
        #Strip off blank space from each line
        for i in range(len(venues)):
            venues[i] = venues[i].strip()
        #Remove empty lines
        for i,venue in enumerate(venues.copy()):
            if venue == '':
                venues.remove(venue)
        
        num_of_venues = len(venues)

        if include_homegrounds:
            if num_of_venues != num_of_teams:
                error_status = 'For homegrounds "Number of venues" should be equal to "Number of teams"'
        else:
            if num_of_venues == 0:
                error_status = 'Enter at least one venue!'

        lbl_numOfVenues.configure(text='Number of venues: '+str(num_of_venues))
    except Exception as e:
        print(e)


def check_dates():
    #Update frame 5 (timeslots)
    if not include_dates:
        btn_51.configure(state='disabled')
        try:
            lbl_ts_stats.configure(text='Time slots are unavailable since \nDates are disabled')
        except:
            load_frame_5()
            lbl_ts_stats.configure(text='Time slots are unavailable since \nDates are disabled')
        return
    else:
        btn_51.configure(state='normal')
        try:
            lbl_ts_stats.configure(text='\n')
        except:
            pass

    
    global min_break
    global error_status

    #Get serial parallel
    if num_of_groups > 1:
        global parallel_groups
        if var_31.get() == 1:
            parallel_groups = False
        elif var_31.get() == 2:
            parallel_groups = True
    
    #----------Get date of first match----------
    for i,entrytype in enumerate(['day','month','year']):
        try:
            if entrytype == 'day':
                var = var_e_31
            elif entrytype == 'month':
                var = var_e_32
            else:
                var = var_e_33
        except:
            return
        empty_input = False
        invalid_input = True
        while 'Checking entry':
            #Input rounds
            value = var.get()
            value = value.strip()
            if value == '':
                invalid_input=False
                empty_input=True
                error_status='"'+ entrytype +'" cannot be empty'
                break
            try:
                value = int(value)
            except:
                validation_msg = '"'+ entrytype +'" should be an integer'
                break
            else:
                invalid_input = False
                break
        
        if empty_input:
            pass
            return
        elif invalid_input:
            error_status=validation_msg
            return
        else:
            error_status=''


def check_matchesPerDay():
    if not include_dates:
        return
    
    global matches_per_day
    global matches_on_sunday
    global error_status

    empty_input = False
    invalid_input = True
    while 'Checking matches per day':
        #Input
        value1 = var_e_wday.get()
        value1 = value1.strip()
        value2 = var_e_sday.get()
        value2 = value2.strip()
        if value1 == '' or value2 == '':
            invalid_input=False
            empty_input=True
            break
        try:
            value1 = int(value1)
            value2 = int(value2)
        except:
            validation_msg = '"number of matches per day" should be an integer'
            break
        if value1 == 0 and value2 == 0:
            validation_msg = 'Set a non zero value for either "matches per weekday"\nor "matches per sunday"'
            break
        else:
            matches_per_day = value1
            matches_on_sunday = value2
            invalid_input = False
            break
    
    if empty_input:
        error_status='"number of matches" cannot be empty'
    elif invalid_input:
        error_status=validation_msg
    else:
        error_status=''

    
def check_firstMatchCostraint():
    global error_status
    try:
        if first_match_constraint:
            if var_team1.get() == var_team2.get():
                error_status = 'Select two different teams for first match!'
    except Exception as e:
        print(e)

    
def check_rounds():
    global rounds
    global error_status

    #----------Check number of rounds----------
    empty_input = False
    invalid_input = True
    while 'Checking number of rounds':
        #Input rounds
        value = var_e_41.get()
        value = value.strip()
        if value == '':
            invalid_input=False
            empty_input=True
            break
        try:
            value = int(value)
        except:
            validation_msg = '"number of rounds" should be an integer'
            break
        if value < 1:
            validation_msg = '"number of rounds" should be from 1-99"'
            break
        else:
            rounds = value
            invalid_input = False
            break
    
    if empty_input:
        error_status='"number of rounds" cannot be empty'
    elif invalid_input:
        error_status=validation_msg
    else:
        error_status=''

    
def check_minbreak():
    global min_break
    global error_status

    #----------Check number of days----------
    empty_input = False
    invalid_input = True
    while 'Checking number of days':
        #Input rounds
        value = var_e_42.get()
        value = value.strip()
        if value == '':
            invalid_input=False
            empty_input=True
            break
        try:
            value = int(value)
        except:
            validation_msg = '"minimum break" should be an integer'
            break
        else:
            min_break = value
            invalid_input = False
            break
    
    if empty_input:
        error_status='"minimum break" cannot be empty'
    elif invalid_input:
        error_status=validation_msg
    else:
        error_status=''


def check_timeslots():
    if (not include_timeslots):
        return

    #Disabling/Normalizing contents
    if not matches_per_day:   
        for child in frame_5_contents:
            try:
                child.configure(state='disable')
            except Exception as e:
                print('frame 5 Error for',type(child))
                print(e)
    else:
        while 'normalizing frame_5_contents':
            try:
                if frame_5_contents[-1].cget('state') == 'normal':
                    break
            except:
                break
            for child in frame_5_contents:
                try:
                    child.configure(state='normal')
                except Exception as e:
                    print('frame 5 Error for',type(child))
                    print(e)
            break

    global error_status
    global timeslots_weekday
    global timeslots_sunday
    global var_numOfTimeslots1
    global var_numOfTimeslots2
        
    timeslots_weekday = []
    timeslots_sunday = []
    
    
    try:
        #----------Input raw text for timeslots1----------
        text1 = e_timeslots1.get('1.0',END)
        slots1 = [line for line in text1.split('\n')]
        #Strip off blank space from each line
        for i in range(len(slots1)):
            slots1[i] = slots1[i].strip()
        #Remove empty lines
        for i,slot in enumerate(slots1.copy()):
            if slot == '':
                slots1.remove(slot)

        denominator = matches_per_day
        if parallel_groups:
            denominator *= num_of_groups
        var_numOfTimeslots1.set('Entries: '+str(len(slots1))+'/'+str(denominator))
        timeslots_weekday = slots1

        #----------Input raw text for timeslots2----------
        text2 = e_timeslots2.get('1.0',END)
        slots2 = [line for line in text2.split('\n')]
        #Strip off blank space from each line
        for i in range(len(slots2)):
            slots2[i] = slots2[i].strip()
        #Remove empty lines
        for i,slot in enumerate(slots2.copy()):
            if slot == '':
                slots2.remove(slot)

        denominator = matches_on_sunday
        if parallel_groups:
            denominator *= num_of_groups
        var_numOfTimeslots2.set('Entries: '+str(len(slots2))+'/'+str(denominator))
        timeslots_sunday = slots2

    except Exception as e:
        pass

    denominator1 = int(var_numOfTimeslots1.get().split('/')[-1])
    denominator2 = int(var_numOfTimeslots2.get().split('/')[-1])
    if len(timeslots_weekday) < denominator1 or len(timeslots_sunday) < denominator2:
        error_status = 'Enter required number of timeslots'
    
    
def set_fmc():
    print('setting')
    teams = groups[0]
    teamnames = teamList(teams)
    i1 = teamnames.index(var_team1.get())
    i2 = teamnames.index(var_team2.get())
    teams[0],teams[i1] = teams[i1],teams[0]
    teams[1],teams[i2] = teams[i2],teams[1]
    groups[0] = teams


def set_timeslots():
    for i,slottext in enumerate(timeslots_weekday):
        hour = slottext[:2]
        mints = slottext[3:5]
        period = slottext[6:8]
        slot = Time(int(hour),int(mints),period)
        timeslots_weekday[i] = slot
    for i,slottext in enumerate(timeslots_sunday):
        hour = slottext[:2]
        mints = slottext[3:5]
        period = slottext[6:8]
        slot = Time(int(hour),int(mints),period)
        timeslots_sunday[i] = slot


def filter_int_inputs():
    #global var_e_11
    filter_int(var_e_11)
    entry_limiter(var_e_11,2)
    try:
        if include_dates:
            filter_int(var_e_31)
            filter_int(var_e_32)
            filter_int(var_e_33)
            entry_limiter(var_e_31,2)
            entry_limiter(var_e_32,2)
            entry_limiter(var_e_33,4)
            filter_int(var_e_wday)
            filter_int(var_e_sday)
            entry_limiter(var_e_wday,1)
            entry_limiter(var_e_sday,1)
    except Exception as e:
        print(e)
    try:
        filter_int(var_e_41)
        filter_int(var_e_42)
        entry_limiter(var_e_41,2)
        entry_limiter(var_e_42,2)
    except Excetion as e:
        print(e)

            
#Main scheduling algorithm
#Takes list of teams 'teams' and minimum break 'min_break'
def schedule(teams,min_break,pflr_index):
    #global include_dates
    global date
    #global matches_per_day
    #global matches_on_sunday
    #global parallel_groups
    #global groups

    pflr = pending_from_last_round[pflr_index]
    
    avail = []
    for team in teams:
        if team.lastplayed(date) >= min_break:
            avail.append(team)
    matches_this_day = []
    if pflr:
        matches_this_day = pflr
    else:
        pflr = None
    matches = []

    #----------Matching Loop----------
    matching = True
    while matching:
        #Find match limit for the day, and set date
        dayOfWeek = get_dayOfWeek(date)
        match_limit = matches_per_day
        if dayOfWeek == 'sunday':
            if matches_on_sunday > 0:
                match_limit = matches_on_sunday
            else:
                #If matches_on_sunday is 0 then go to monday
                #also increase date of current match
                if parallel_groups:
                    date.inc_day()
                else:
                    for i in range(len(groups)):
                        date.inc_day()
                        match[2].inc_day()
                match_limit = matches_per_day

        #Matching two available teams        
        match = None
        if len(avail) > 0:
            for i in range(len(avail)-1):
                if match:
                    break
                for j in range(i+1,len(avail)):
                    a,b = avail[i],avail[j]
                    if a in b.teamsplayed:
                        continue
                    avail.remove(a)
                    avail.remove(b)
                    a.teamsplayed.append(b)
                    b.teamsplayed.append(a)
                    a.lastdate = deepcopy(date)
                    b.lastdate = deepcopy(date)
                    d = date.get_date()
                    match = [a,b,Date(d)] 
                    break
        if not match:
            break
        else:
            matches_this_day.append(match)
            matches.append(match)

        
        #Increment date based on parallel_groups
        if len(matches_this_day) == match_limit:
            if parallel_groups:
                date.inc_day()
            else:
                for i in range(len(groups)):
                    date.inc_day()
            matches_this_day = []

        #Prepare avail for next iteration based on which teams got enough rest
        avail = []
        for team in teams:
            if team.lastplayed(date) >= min_break:
                avail.append(team)

        random.shuffle(avail)
        sortTeams(avail)

        
    #----------End WHILE LOOP----------

    #If while loop terminated prematurely, return False
    if len(matches) < expectedLen(len(teams)):
        return False

    print('Total:',len(matches))

    print('!!!!!')
    print('')
    print(date.get_date())
    print('matches_this_day:',matches_this_day)
    print('match_limit:',match_limit)
    if len(matches_this_day) < match_limit:
        pflr = matches_this_day
        pending_from_last_round[pflr_index] = pflr
    return matches


def schedule_allGroups():
    global grouped_matches_set
    global date
    global pending_from_last_round


    #Set try_limit to number of teams in each group, but set to 10 if it's lesser
    try_limit = len(groups[0]) * 2
    if try_limit < 10:
        try_limit = 10
    elif try_limit > 30:
        try_limit = 30
        
    grouped_matches_set = []
    date_prevGrp = deepcopy(date)
    #Generate schedule for each group
    pending_from_last_round = []
    for groupindex,group in enumerate(groups):
        pending_from_last_round.append([])
        grouped_matches_set.append([])
        print('Group',chr(65+groupindex))
        print('******************************')
        if not parallel_groups and groupindex != 0:
            date = date_prevGrp
            date.inc_day()
            date_prevGrp = deepcopy(date)
        elif parallel_groups:
            date = date_prevGrp
            date_prevGrp = deepcopy(date)
        #Generate schedule for each round

        for r in range(1,rounds+1): 
            tries = 0
            #Shuffle teams based on first match constraint
            if not (first_match_constraint and groupindex == 0 and r == 1):
                random.shuffle(group)
               
            scheduled = False
            #Keep looping till schedule generated or failed inspite of trying a lot!
            while not scheduled:
                tries += 1
                resetTeams(group)
                date_backup = deepcopy(date)
                group_backup = deepcopy(group)
                ##print([[t.name,t.lastplayed] for t in group])
                scheduled = schedule(group,min_break,pflr_index=groupindex)
                ##print('After scheduling')
                ##print([[t.name,t.lastplayed] for t in group])
                if not scheduled:
                    date = date_backup
                    group = group_backup
                else:
                    grouped_matches_set[groupindex] += scheduled
                    print('success')
                #Since group is now a separate copy of group, we need to update it in groups
                groups[groupindex] = group
                if tries > try_limit:
                    error_msg = '"minimum break" or "matches per day" is too high'
                    print(error_msg)
                    return error_msg
    return 0


def compile_matchlist():
    global all_matches
    print('compiling matches')
    print(len(grouped_matches_set))
    
    #Compile the scheduled matches in a list
    all_matches = []
    for groupindex,grouped_matches in enumerate(grouped_matches_set):
        for m in grouped_matches:
            match = Match(m[0],m[1],chr(65+groupindex),m[2],None)
            all_matches.append(match)


def update_venues():
    print('in update_venues()')
    #Include Venues
    if include_homegrounds:
        print('in if block')
        matches_per_team = (len(groups[0])-1) * rounds
        #For each team find pairs with opponents
        for group in groups:
            for team in group:
                pairs = []
                for opp in team.teamsplayed:
                    pairs.append((team,opp))
                for pair in pairs.copy():
                    pair_matches = []
                    pair_already_done = False
                    for match in all_matches:
                        if pair[0].name in match.getTeamNames() and pair[1].name in match.getTeamNames():
                            if match.venue != None:
                                pairs.remove(pair)
                                pair_already_done = True
                                break
                            pair_matches.append(match)
                    if pair_already_done:
                        continue
                    #Alternate home and away matches for the pair matches
                    avail_venues = [pair[0].homeground,pair[1].homeground]
                    random.shuffle(avail_venues)
                    v_index = 0
                    for match in pair_matches:
                        match.venue = avail_venues[v_index]
                        print(match.team1.name,'vs',match.team2.name,match.venue,avail_venues,v_index)
                        v_index = -(v_index - 1)
    elif include_venues:
        sortMatches(all_matches,'date')
        venue_index = 0
        for match in all_matches:
            match.venue = venues[venue_index]
            venue_index += 1
            if venue_index == len(venues):
                venue_index = 0


def update_timeslots():
    print('in update_timeslots()')
    #Give time slots
    sortMatches(all_matches,'date')
    if include_timeslots:
        previous_day = date.firstdate
        slot_index = 0
        for m in all_matches:
            match_date = m.date.get_date()
            if include_timeslots:
                isSunday = get_dayOfWeek(m.date) == 'sunday'
                time_slot = None
                if previous_day == match_date:
                    if isSunday:
                        print('slot_index:',slot_index)
                        try:
                            time_slot = timeslots_sunday[slot_index]
                        except:
                            print(match_date)
                            print([[teamList([m.team1,m.team2]),m.date.get_date()] for m in all_matches])
                            return
                    else:
                        time_slot = timeslots_weekday[slot_index]
                    slot_index += 1
                else:
                    previous_day = match_date
                    slot_index = 0
                    if isSunday:
                        time_slot = timeslots_sunday[slot_index]
                    else:
                        time_slot = timeslots_weekday[slot_index]
                    slot_index += 1
            m.time = time_slot
    print('exiting update_venues()')


def print_allmatches():
    #display all matches
    for m_index,m in enumerate(all_matches):
        m_date = m.date.get_date()
        if include_timeslots:
            m_time = m.time.getTimeString()
        else:
            m_time = None
        m_num  = m_index + 1
        m_t1   = m.team1.name
        m_t2   = m.team2.name
        m_venue= m.venue
        print(m_date,m_time,'  MATCH',m_num,m_t1,'vs',m_t2,'at',m_venue)


def print_teammatches():
    #display matches per team
    for team in all_teams:
        print('----------------------')
        print(team.name)
        print('----------------------')
        for m_index,m in enumerate(all_matches):
            m_date = m.date.get_date()
            if include_timeslots:
                m_time = m.time.getTimeString()
            else:
                m_time = None
            m_num  = m_index + 1
            m_t1   = m.team1.name
            m_t2   = m.team2.name
            m_venue= m.venue
            if team.name in m.getTeamNames():
                home = 'Home' if team.homeground == m_venue else 'Away'
                print(m_date,m_time,'  MATCH',m_num,m_t1,'vs',m_t2,'at',m_venue,home)



    


    






























#*****************************************************
#*****************RECURRING FUNCTION******************

def update():
    global error_status
    
    error_status = ''

    if editing_teamsAndVenues:
        btn_left.configure(state='disable')
        btn_schedule.configure(state='disable')
        check_numOfTeams()
        if error_status=='':
            check_numOfGroups()
        if error_status=='' and num_of_teams>0:
            process_teams()
        if error_status=='':
            check_venues()
    else:
        check_dates()
        if error_status=='':
            check_matchesPerDay()
        if error_status=='':
            check_firstMatchCostraint()
        if error_status=='':
            check_rounds()
        if error_status=='':
            check_minbreak()
        if error_status=='':
            check_timeslots()
        else:
            try:
                if frame_5_contents[-1].cget('state') == 'normal':
                    for child in frame_5_contents:
                        try:
                            child.configure(state='disable')
                        except Exception as e:
                            print('frame 5 Error for',type(child))
                            print(e)
            except:
                pass
        if error_status=='':
            btn_schedule.configure(state='normal')
        else:
            btn_schedule.configure(state='disable')
            
            
    if error_status=='':
        btn_left.configure(state='normal')
    
    lbl_status.configure(text=error_status,fg='Yellow')

    filter_int_inputs()
    
    w.after(1,update)
























#**************************************************
#*****************EVENT FUNCTIONS******************

#On pressing: left_btn (a)
def lock_teamsAndVenues():
    
    #Reset the variables for team1 and team2 of first match constraint
    global var_team1
    global var_team2
    global lbl_3_height
    try:
        var_team1.set(teamList(groups[0])[0])
        var_team2.set(teamList(groups[0])[1])
    except:
        print('Cant set var_team1 and var_team2')

    #Destroy the subframe frame_4_firstmatch
    try:
        print('Destroying frame_4_firstmatch')
        frame_4_firstmatch.destroy()
    except Exception as e:
        print(e)
        print()

    #Load new frame_4_firstmatch subframe
    load_frame_4_fmc_a()

    #if only one group,destroy serial parallel contents
    if num_of_groups == 1:
        try:
            for i in range(len(serialparallel_contents)-1,-1,-1):
                try:
                    serialparallel_contents[i].destroy()
                    serialparallel_contents.pop(i)
                except:
                    print('serialparallel destroy child error for',type(child))
        except Exception as e:
            print('Serial parallel error:',e)
        #Height adjustment for frame 3
        try:
            lbl_3_height.destroy()
        except Exception as e:
            print(e)
        try:
            print('creating...')
            lbl_3_height = Label(frame_3e,text='',bg=widgetColor('frame_3'))
            lbl_3_height.grid(row=20,pady=46)
        except Exception as e:
            print(e)
    #if more than one group, load serial parallel contents
    else:
        try:
            serialparallel_contents == []
            if serialparallel_contents == []:
                load_frame_3_spc()
        except:
            pass
        try:
            print('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
            lbl_3_height.destroy()
        except Exception as e:
            print(e)
        
    #Disable contents of frame 1 and 2
    for child in frame_1_contents + frame_2_contents:
        try:
            child.configure(state='disable')
        except:
            print('frame 1,2 Error for',type(child))

    #Normalize contents of frame 3 and 4
    for child in frame_3_contents + frame_4_contents + frame_5_contents:
        try:
            child.configure(state='normal')
        except Exception as e:
            print('frame 3,4,5 Error for',type(child))
            print(e)
        
    #Load leftButton_b, 'Edit teams and venues'
    load_leftButton_b()



#On pressing: left_btn (b)
def edit_teamsAndVenues(): 

    for child in frame_1_contents + frame_2_contents:
        try:
            child.configure(state='normal')
        except:
            print('frame 1,2 Error for',type(child))

    
    for child in frame_3_contents + frame_4_contents + frame_5_contents:
        try:
            child.configure(state='disable')
        except Exception as e:
            print('frame 3,4,5 Error for',type(child))
            print(e)

    load_leftButton_a()



#On pressing: btn_add_timeslot [1&2]
def add_timeslot(daytype):
    if daytype =='weekday':
        entry_box = e_timeslots1
        hour = var_hour1.get()
        mints = var_mints1.get()
        period = var_period1.get()
    elif daytype =='sunday':
        entry_box = e_timeslots2
        hour = var_hour2.get()
        mints = var_mints2.get()
        period = var_period2.get()

    line = hour+':'+mints+' '+period+'\n'
    entry_box.configure(state='normal')
    #Create list of time slots from entry box
    text = entry_box.get('1.0',END)
    slots = (text+line).split('\n')
    #Remove blank lines from slots and create new text
    newtext=''
    
    for slot in slots.copy():
        if slot == '':
            slots.remove(slot)
        else:
            newtext += slot +'\n'

    print(slots)

    entry_box.delete('1.0',END)
    entry_box.insert('1.0',newtext)
    entry_box.configure(state='disable')
        


#On pressing: btn_rem_timeslot [1&2]
def remove_timeslot(daytype):
    if daytype =='weekday':
        entry_box = e_timeslots1
    elif daytype =='sunday':
        entry_box = e_timeslots2

    entry_box.configure(state='normal')
    #Create list of time slots from entry box
    text = entry_box.get('1.0',END)
    slots = (text).split('\n')
    #Remove blank lines from slots
    for slot in slots.copy():
        if slot == '':
            slots.remove(slot)
    #Create new text having the last slot removed
    newslots=slots[:-1]
    newtext=''
    for slot in newslots:
        newtext += slot+'\n'

    print(newslots)

    entry_box.delete('1.0',END)
    entry_box.insert('1.0',newtext)
    entry_box.configure(state='disable')



#On pressing: btn_schedule
def executeSchedule():
    global date

    #Set last played date for all teams as [1,1,1]
    set_teamlastdate()

    #Set homeground for all teams if including homegrounds
    if include_homegrounds:
        set_homegrounds()

    #Create date object by inputing date specified by user
    if include_dates:
        date = Date([int(var_e_33.get()),int(var_e_32.get()),int(var_e_31.get())])
    else:
        date = Date([2022,1,1])
        global matches_per_day
        global matches_on_sunday
        matches_per_day = 1
        matches_on_sunday = 1

    #Set First Match Constraint
    if first_match_constraint:
        set_fmc()

    #Set Timeslots
    if include_timeslots:
        set_timeslots()

    print(f'''num_of_teams = {num_of_teams}
num_of_groups = {num_of_groups}

grouplen = {grouplen}
groups = {groups}
all_teams = {all_teams}

min_break = {min_break}
try_limit = {try_limit}
rounds = {rounds}
first_match_constraint = {first_match_constraint}
include_dates = {include_dates}
include_timeslots = {include_timeslots}
include_venues = {include_venues}
include_homegrounds = {include_homegrounds}
parallel_groups = {parallel_groups}

date = {date}
matches_per_day = {matches_per_day}
matches_on_sunday = {matches_on_sunday}

match_duration = {match_duration}
timeslots_weekday = {timeslots_weekday}
timeslots_sunday = {timeslots_sunday}

venues = {venues}

''')
    #Schedule all groups, if an error occurs store it in a variable
    error_while_scheduling = schedule_allGroups()
    if error_while_scheduling:
        popup_error('Error while scheduling',error_while_scheduling)
        return
    
    #Compile list of matches
    compile_matchlist()
    
    #Update venues to the matches
    update_venues()
    
    try:
        error_timeslots = update_timeslots()
    except Exception as e:
        error_timeslots = e
    
    #Update timeslots to the matches
    if error_timeslots:
        popup_error('Timeslots error',str(error_timeslots))
        return

    schedule_object = Schedule(all_teams,all_matches,include_venues,include_homegrounds,include_dates,parallel_groups,include_timeslots)
    #Output
    print(schedule_object.get_grouplen())
    schedule_object.print_allmatches()
    schedule_object.print_teammatches()
    schedule_object.xl_output('demo')
    '''
    print_allmatches()
    print_teammatches()
    print('done')
    '''

































#****************************************************
#********************GUI FUNCTIONS*******************

def colorize_widget(widget):
    if widget == 'frame_1':
        for child in frame_1_contents:
            try:
                print('coloring')
                child.configure(background=widgetColor('frame_1'))
            except Exception as e:
                print(e)
    elif widget == 'frame_2':
        for child in frame_2_contents:
            try:
                print('coloring')
                child.configure(background=widgetColor('frame_2'))
            except Exception as e:
                print(e)
    elif widget == 'frame_3':
        for child in frame_3_contents:
            try:
                child.configure(background=widgetColor('frame_3'))
            except:
                print('Error for',type(child))
    elif widget == 'frame_4':
        for child in frame_4_contents:
            try:
                child.configure(background=widgetColor('frame_4'))
            except:
                print('Error for',type(child))
    elif widget == 'frame_5':
        for child in frame_5_contents:
            try:
                child.configure(background=widgetColor('frame_5'))
            except:
                print('Error for',type(child))
            
    
def load_mainframe(frame_name,destroy=False):
    if destroy:
        try:
            current_frame.destroy()
        except Exception as e:
            print(e)
            print()
    load_framename = eval('load_frame_'+frame_name)
    load_framename()


#fmainInput
def load_frame_mainInput():
    global frame_mainInput

    #Create frame_mainInput
    #frame_mainInput = LabelFrame(w,width=500, height = 500,padx = 10, pady = 10)
    frame_mainInput = LabelFrame(w,borderwidth=4,padx = 10, pady = 10)
    frame_mainInput.pack(padx = 20, pady = 20)
    frame_mainInput.place(width=1100, height = 560)
    #frame_mainInput.configure(bg=widgetColor('frame_mainInput'))
    frame_mainInput.configure(bg='grey70')

    #Load child frames
    load_frame_1()
    load_frame_2()
    load_frame_3()
    load_frame_4()
    load_frame_5()

    print('frame 4 contents:\n',len(frame_4_contents))

    for child in frame_3_contents + frame_4_contents + frame_5_contents:
        try:
            child.configure(state='disable')
        except:
            print('Error for',type(child))

    
#f1 Teams and Groups
def load_frame_1():
    global frame_1
    global e_11
    global var_e_11
    global e_teams
    global lbl_numOfTeamsAndGroups
    global frame_1_contents
    
    #Create frame_1
    frame_1 = LabelFrame(frame_mainInput,text='Teams and Groups',padx = 10, pady = 10 )
    f = frame_1
    f.configure(background=widgetColor('frame_1'))

    #define contents of frame_1
    lbl_11 = Label(f,text='Number of Groups: ',pady=5,padx=5)
    var_e_11 =StringVar()
    var_e_11.set('1')
    e_11 = Entry(f,width=2,textvariable=var_e_11)
    lbl_empty = Label(f,text='')
    e_teams = ScrolledText(f,width=20,height = 20)
    lbl_numOfTeamsAndGroups = Label(f,text='',pady=5)
    lbl_empty2 = Label(f,text=' '*60,pady=1)

    frame_1_contents = [lbl_11,e_11,lbl_empty,e_teams,lbl_numOfTeamsAndGroups,lbl_empty2]
    
    #position contents of frame_1
    lbl_11.grid(sticky=W)
    e_11.grid(row=0,column=1,sticky=W)
    lbl_empty.grid()
    e_teams.grid(row=3,columnspan=2)
    lbl_numOfTeamsAndGroups.grid(row=4,sticky=E)

    #Empty Label at end to fix width of parent frame
    lbl_empty2.grid(row=6,columnspan=3)

    colorize_widget('frame_1')
    
    f.grid(rowspan=2,padx = 10, pady = 10, sticky=W+N)

    
#f2 Venues (disabled)
def load_frame_2():
    global frame_2
    global btn_21
    global include_venues
    global include_homegrounds
    global frame_2_contents

    include_venues = False
    include_homegrounds = False

    #Destroy frame_2e before loading frame_2 
    try:
        print('destroying frame_2e')
        frame_2e.destroy()
    except Exception as e:
        print(e)
        print()
    
    #Create frame_2
    frame_2 = LabelFrame(frame_mainInput,text='Venues',width=100,padx = 10, pady = 10)
    f = frame_2
    f.configure(background=widgetColor('frame_2'))
        
    #Create button 'Add Venues'
    btn_21 = Button(f, text='Add Venues', command =load_frame_2e,cursor='hand2')
    btn_21.grid(row=0,columnspan=2)

    #Empty Label at end to fix width of parent frame
    lbl_empty=Label(f,text=' '*58)
    lbl_empty.grid(row=1,pady=196)

    frame_2_contents = [btn_21,lbl_empty]

    colorize_widget('frame_2')
    
    f.grid(row=0,column=1,rowspan=2,padx = 10, pady = 10,sticky=W+N)


#f2a Venues (enabled)
def load_frame_2e():
    global frame_2e
    global var_21
    global btn_21
    global e_venues
    global lbl_numOfVenues
    global include_venues
    global include_homegrounds
    global frame_2_contents

    include_venues = True
    include_homegrounds = False
    
    #Destroy frame_2 before loading frame_2e 
    try:
        print('destroying frame_2')
        frame_2.destroy()
    except Exception as e:
        print(e)
        print()
    
    #Create frame_2e
    frame_2e = LabelFrame(frame_mainInput,text='Venues',padx = 10, pady = 10)
    f = frame_2e
    f.configure(background=widgetColor('frame_2'))

    #Destroy button 'Add Venues' and create 'Cancel Venues'
    try:
        print('destroying btn_21')
        btn_21.destroy()
    except Exception as e:
        print(e)
        print()
    btn_21 = Button(f,text='Cancel Venues', command =load_frame_2,cursor='hand2')
    btn_21.grid()

    #Give a line break and create a ScrolledText for venues
    lbl_empty=Label(f,text='',pady=3)
    lbl_empty.grid()
    e_venues = ScrolledText(f,width=20,height=20)
    e_venues.grid()
    
    #Input whether venues are homegrounds
    var_21 = IntVar()
    cb_21 = Checkbutton(f,text='Venues are homegrounds',variable = var_21,cursor='hand2',padx=5,pady=5)      
    cb_21.grid()

    #Number of venues
    lbl_numOfVenues = Label(f,text='',pady=5)
    lbl_numOfVenues.grid(pady=1)

    #Empty Label at end to fix width of parent frame
    lbl_empty2 = Label(f,text=' '*60)
    #lbl_empty2.grid(row=6)
    
    frame_2_contents = [btn_21,lbl_empty,cb_21,e_venues,lbl_numOfVenues,lbl_empty2]

    colorize_widget('frame_2')
    
    f.grid(row=0,column=1,rowspan=2,padx = 10, pady = 10,sticky=W+N)

    
#f3 Dates (disabled)
def load_frame_3():
    global frame_3
    global include_dates
    global frame_3_contents

    include_dates = False

    frame_3_contents = []
    
    #Destroy frame_3e before loading frame_3 
    try:
        print('destroying frame_3e')
        frame_3e.destroy()
    except Exception as e:
        print(e)
        print()
    
    #Create frame_3
    frame_3 = LabelFrame(frame_mainInput,text='Dates',padx = 10, pady = 10)
    f = frame_3
    f.configure(background=widgetColor('frame_3'))
    
    #Create button 'Add Dates'
    btn_31 = Button(f,text='Add Dates', command =load_frame_3e,cursor='hand2')
    btn_31.grid(row=0)
    lbl_empty1 = Label(f,text=' '*33,bg=widgetColor('frame_3'))
    lbl_empty2 = Label(f,text=' '*33,bg=widgetColor('frame_3'))
    #lbl_empty1.grid(row=0,column=0)
    #lbl_empty2.grid(row=0,column=2)
    Label(f,text='',padx=136.8,pady=97,bg=widgetColor('frame_3')).grid()
    

    frame_3_contents = [btn_31,lbl_empty1,lbl_empty2]

    colorize_widget('frame_3')
    
    f.grid(row=0,column=2,padx = 10, pady = 10,sticky=W+N)


#f3e Dates (enabled)
def load_frame_3e():
    global frame_3e
    global frame_3_contents
    global frame_3_base_contents
    global btn_31
    global include_dates
    global include_timeslots
    global parallel_groups
    global var_e_31
    global var_e_32
    global var_e_33
    global var_31
    global var_e_wday
    global var_e_sday
    global btn_ts

    include_dates = True
    parallel_groups = False
    
    #Destroy frame_3 before loading frame_3e
    try:
        print('destroying frame_3')
        frame_3.destroy()
    except Exception as e:
        print(e)
        print()
    
    #Create frame_3e
    frame_3e = LabelFrame(frame_mainInput,text='Dates',padx = 10, pady = 10)
    f = frame_3e
    f.configure(background=widgetColor('frame_3'))

    #-------------------
    #Destroy button 'Add Dates' and create 'Cancel Dates'
    try:
        print('destroying btn_31')
        btn_31.destroy()
    except Exception as e:
        print(e)
    btn_31 = Button(f,text='Cancel Dates', command =load_frame_3,cursor='hand2')
    btn_31.grid(columnspan=4)
    lbl_empty = Label(f,text='')
    lbl_empty.grid(columnspan=4)
    #-------------------
    #Update global variable include_dates
    include_dates = True
    #-------------------
    #Starting date i.e date of first match
    lbl_34 = Label(f,text='Date of first match: (dd/mm/yyyy)',padx=5)
    var_e_31 = StringVar()
    var_e_32 = StringVar()
    var_e_33 = StringVar()
    var_e_31.set('1')
    var_e_32.set('1')
    var_e_33.set('2022')
    e_31 = Entry(f,width=2,textvariable=var_e_31)
    e_32 = Entry(f,width=2,textvariable=var_e_32)
    e_33 = Entry(f,width=4,textvariable=var_e_33)
    
    lbl_34.grid(row=7,column=0,sticky=W)
    e_31.grid(row=7,column=1)
    e_32.grid(row=7,column=2)
    e_33.grid(row=7,column=3)
    #-------------------
    #Matches per week day
    lbl_wday = Label(f,text='Matches on week days for each group:',pady=5,padx=5)
    var_e_wday =StringVar()
    var_e_wday.set('1')
    e_wday = Entry(f,width=2,textvariable=var_e_wday)
    
    lbl_wday.grid(row=8,column=0,sticky=W)
    e_wday.grid(row=8,column=1)
    #-------------------
    #Matches per sunday
    lbl_sday = Label(f,text='Matches on Sundays for each group:',pady=5,padx=5)
    var_e_sday =StringVar()
    var_e_sday.set('1')
    e_sday = Entry(f,width=2,textvariable=var_e_sday)
    
    lbl_sday.grid(row=9,column=0,sticky=W)
    e_sday.grid(row=9,column=1)
    #-------------------
    frame_3_base_contents = [btn_31,lbl_empty,lbl_34,e_31,e_32,e_33,lbl_wday,e_wday,lbl_sday,e_sday]
    frame_3_contents = frame_3_base_contents
    #-------------------
    #Serial Parallel contents (if num_of_groups == 1, then spc contents won't be placed in the window)
    var_31 = IntVar()
    var_31.set(1)
    load_frame_3_spc()
    #-------------------
    colorize_widget('frame_3')
    
    f.grid(row=0,column=2,padx = 10, pady = 10,sticky=W+N)


def load_frame_3_spc():
    global serialparallel_contents
    global frame_3_contents
    global lbl_3_height

    f = frame_3e
    
    #Height adjust
    try:
        lbl_3_height.destroy()
    except Exception as e:
        print(e)
    print('creating...')
    lbl_3_height = Label(f,text='',bg=widgetColor('frame_3'))
    lbl_3_height.grid(row=20,pady=46)
    
    serialparallel_contents = []
    lbl_31 = Label(f,text='How many groups play on the same day?')
    options_31 = ['Serial Group Matching','Parallel Group Matching']
    
    rb_31 = Radiobutton(f,text=options_31[0],variable=var_31,value=1,cursor='hand2')
    rb_32 = Radiobutton(f,text=options_31[1],variable=var_31,value=2,cursor='hand2')
    lbl_32 = Label(f,text='(only one group plays per day)')
    lbl_33 = Label(f,text='(each group plays on the same day)')
    #Only ask about serial parallel if num_of_groups > 1
    if num_of_groups>1:
        #note: the list serialparallel_contents is empty if and only if num_of_groups is 1
        #it's filled otherwise
        #this info about the list is used in the program when calling the function load_frame_3_spc
        serialparallel_contents += [lbl_31,rb_31,lbl_32,rb_32,lbl_33]
        frame_3_contents = frame_3_base_contents + serialparallel_contents
        print('!!!!!!!!!!!!!')
        colorize_widget('frame_3')
        lbl_31.grid(row=2,columnspan=4)
        rb_31.grid(row=3,sticky=W,columnspan=4)
        lbl_32.grid(row=4,sticky=E,columnspan=4)
        rb_32.grid(row=5,sticky=W,columnspan=4)
        lbl_33.grid(row=6,sticky=E,columnspan=4)
        try:
            lbl_3_height.destroy()
        except Exception as e:
            print(e)
    

#f4 Parameters
def load_frame_4():
    global frame_4
    global frame_4_contents
    global var_e_41
    global var_e_42
    
    frame_4_contents = []
    
    #Create frame_4
    frame_4 = LabelFrame(frame_mainInput,text='Parameters',padx = 10, pady = 10)
    f = frame_4
    f.configure(background=widgetColor('frame_4'))

    lbl_41 = Label(f,text='Number of Rounds',pady=5,padx=5)
    var_e_41 =StringVar()
    var_e_41.set('1')
    e_41 = Entry(f,width=2,textvariable=var_e_41)
    lbl_42 = Label(f,text='Minimum break for teams in days',padx=5)
    lbl_43 = Label(f,text='(i.e rest for each team before\nconsecutive matches)')
    var_e_42 =StringVar()
    var_e_42.set('1')
    e_42 = Entry(f,width=2,textvariable=var_e_42)

    lbl_41.grid(row=1,sticky=W)
    e_41.grid(row=1,column=1,sticky=W)
    lbl_42.grid(row=2,sticky=W)
    lbl_43.grid(row=3,columnspan=2,sticky=W)
    e_42.grid(row=2,column=1,sticky=W)

    frame_4_contents = [lbl_41,e_41,lbl_42,lbl_43,e_42]

    #First Match Constraint
    load_frame_4_fmc_a()

    #exception handling: two teams shouldn't be same

    colorize_widget('frame_4')

    f.grid(row=1,column=2,columnspan=2,padx = 10, pady = 10,sticky=W+N)


#f4_fmc_a first_match_constraint (disabled)
def load_frame_4_fmc_a():
    global frame_4_firstmatch
    global btn_41
    global first_match_constraint

    first_match_constraint = False

    try:
        print('Destroying frame_4_firstmatch')
        frame_4_firstmatch.destroy()
    except Exception as e:
        print(e)
        print()
    
    frame_4_firstmatch = LabelFrame(frame_4,padx = 10, pady = 10)
    f2 = frame_4_firstmatch
    f2.configure(background=widgetColor('frame_4'))
    
    btn_41 = Button(f2,text='Enable First Match Constraint',command=load_frame_4_fmc_b,cursor='hand2')
    btn_41.grid(columnspan=3,sticky=W,padx=44,pady=21)

    try:
        for i in range(len(frame_4_contents)-1,4.-1):
            frame_4_contents[i].pop()
    except Exception as e:
        print(e)
    finally:
        frame_4_contents.append(btn_41)

    colorize_widget('frame_4')

    f2.grid(row=0,columnspan = 3,sticky=W)


#f4_fmc_b first_match_constraint (enabled)
def load_frame_4_fmc_b():
    global frame_4_contents
    global var_inc_fmc
    global btn_41
    global frame_4_firstmatch
    global first_match_constraint
    global var_team1
    global var_team2

    first_match_constraint = True
 
    try:
        print('Destroying frame_4_firstmatch')
        frame_4_firstmatch.destroy()
    except Exception as e:
        print(e)
        print()
        
    frame_4_firstmatch = LabelFrame(frame_4,padx = 10, pady = 10)
    f2 = frame_4_firstmatch
    f2.configure(background=widgetColor('frame_4'))

    btn_41 = Button(f2,text='Disable First Match Constraint ',command=load_frame_4_fmc_a,cursor='hand2')
    btn_41.grid(columnspan=3,padx=41,pady=5)
    
    #team_names = []
    #groups = [['a1','a2','a3'],['b1','b2','b3'],['c1','c2','c3']]
    #for team in groups[0]:
        #pass
    team_names = teamList(groups[0]) #!!! team.name !!!
    
    var_team1 = StringVar()
    var_team1.set(team_names[0])
    drop_team1 = OptionMenu(f2, var_team1, *team_names)      
    drop_team1.grid(row=1,column=0,sticky=E)
    
    Label(f2,text='vs').grid(row=1,column=1)
    
    var_team2 = StringVar()
    var_team2.set(team_names[1])
    drop_team2 = OptionMenu(f2, var_team2, *team_names)      
    drop_team2.grid(row=1,column=2,sticky=W)

    try:
        for i in range(len(frame_4_contents)-1,4.-1):
            frame_4_contents[i].pop()
    except Exception as e:
        print(e)
    finally:
        frame_4_contents += [btn_41,drop_team1,drop_team2]
        
    colorize_widget('frame_4')

    f2.grid(row=0,columnspan = 3,sticky=W)


#f5 Timeslots (disabled)
def load_frame_5():
    global frame_5
    global btn_51
    global lbl_ts_stats
    global include_timeslots
    global frame_5_contents

    include_timeslots = False

    #Destroy frame_5e before loading frame_5 
    try:
        print('destroying frame_5e')
        frame_5e.destroy()
    except Exception as e:
        print(e)
        print()
    
    #Create frame_2
    frame_5 = LabelFrame(frame_mainInput,text='Timeslots',width=100,padx = 10, pady = 10)
    f = frame_5
    f.configure(background=widgetColor('frame_5'))
        
    #Create button 'Add Venues'
    btn_51 = Button(f, text='Add Timeslots', command =load_frame_5e,cursor='hand2')
    btn_51.grid(row=0,columnspan=2,padx=58)

    #Empty Label at end to fix width of parent frame
    lbl_ts_stats=Label(f,text='\n')#*58)
    lbl_ts_stats.grid(row=1)

    #Height adjust
    Label(f,text='',bg=widgetColor('frame_5')).grid(pady=196)

    frame_5_contents = [btn_51,lbl_ts_stats]

    colorize_widget('frame_5')
    
    f.grid(row=0,column=3,rowspan=2,padx = 10, pady = 10,sticky=W+N)


#f5e Timeslots (enabled)
def load_frame_5e():
    global frame_5e
    global btn_51
    global e_timeslots1
    global e_timeslots2
    global var_numOfTimeslots1
    global var_numOfTimeslots2
    global var_hour1
    global var_mints1
    global var_period1
    global var_hour2
    global var_mints2
    global var_period2
    global include_timeslots
    global frame_5_contents

    include_timeslots = True

    #Destroy frame_5 before loading frame_5e 
    try:
        print('destroying frame_5')
        frame_5.destroy()
    except Exception as e:
        print(e)
        print()
    
    #Create frame_5a
    frame_5e = LabelFrame(frame_mainInput,text='Timeslots',padx = 10, pady = 10)
    f = frame_5e
    f.configure(background=widgetColor('frame_5'))

    #Destroy button 'Add Timeslots' and create 'Cancel Timeslots'
    try:
        print('destroying btn_51')
        btn_51.destroy()
    except Exception as e:
        print(e)
        print()
    btn_51 = Button(f,text='Cancel Timeslots', command =load_frame_5,cursor='hand2')
    btn_51.grid(row=0,columnspan=4,pady=8,sticky=N)

    #Give a line break 
    lbl_empty=Label(f,text='',pady=3)
    #lbl_empty.grid(row=1,columnspan=6)

    #Define lists of hours, minutes and periods
    hours = ['0'+str(i) if i<10 else str(i) for i in range(1,13)]
    mints = ['0'+str(i) if i<10 else str(i) for i in range(0,56,5)]
    periods=['AM','PM']
    
    #Create Label, Time Picker and ScrolledText for timeslots1
    lbl_timeslots1 = Label(f,text='Week Day Time Slots:')
    lbl_timeslots1.grid(row=2,columnspan=6)

    var_hour1 = StringVar()
    var_hour1.set('12')
    var_mints1 = StringVar()
    var_mints1.set('00')
    var_period1 = StringVar()
    var_period1.set('AM')
    drop_hour1 = OptionMenu(f, var_hour1, *hours)      
    drop_hour1.grid(row=3,column=0)
    lbl_colon1 = Label(f,text=':')
    lbl_colon1.grid(row=3,column=1)
    drop_mints1 = OptionMenu(f, var_mints1, *mints)      
    drop_mints1.grid(row=3,column=2)
    drop_period1 = OptionMenu(f, var_period1, *periods)      
    drop_period1.grid(row=3,column=3)
    btn_add_timeslot1 = Button(f,text='Add Timeslot',command=lambda: add_timeslot('weekday'))
    btn_add_timeslot1.grid(row=4,column=0,columnspan=2,pady=5)
    btn_rem_timeslot1 = Button(f,text='Remove Timeslot',command=lambda: remove_timeslot('weekday'))
    btn_rem_timeslot1.grid(row=4,column=2,columnspan=2,sticky=E)
    timepicker_contents1 = [drop_hour1,lbl_colon1,drop_mints1,drop_period1,btn_add_timeslot1,btn_rem_timeslot1]
    
    e_timeslots1 = ScrolledText(f,width=10,height=6,bg=widgetColor('frame_5'))
    e_timeslots1.grid(row=5,columnspan=6)
    e_timeslots1.configure(state='disable')

    #Number of timeslots entered for week days
    var_numOfTimeslots1 = StringVar()
    var_numOfTimeslots1.set('Entries: 0/'+str(matches_per_day))
    lbl_numOfTimeslots1 = Label(f,textvariable=var_numOfTimeslots1,pady=5)
    lbl_numOfTimeslots1.grid(row=6,columnspan=6)

    #Create Label and ScrolledText for timeslots1
    lbl_timeslots2 = Label(f,text='Sunday Time Slots:')
    lbl_timeslots2.grid(row=7,columnspan=6)
    
    var_hour2 = StringVar()
    var_hour2.set('12')
    var_mints2 = StringVar()
    var_mints2.set('00')
    var_period2 = StringVar()
    var_period2.set('AM')
    drop_hour2 = OptionMenu(f, var_hour2, *hours)      
    drop_hour2.grid(row=8,column=0)
    lbl_colon2 = Label(f,text=':')
    lbl_colon2.grid(row=8,column=1)
    drop_mints2 = OptionMenu(f, var_mints2, *mints)      
    drop_mints2.grid(row=8,column=2)
    drop_period2 = OptionMenu(f, var_period2, *periods)      
    drop_period2.grid(row=8,column=3)
    btn_add_timeslot2 = Button(f,text='Add Timeslot',command=lambda: add_timeslot('sunday'))
    btn_add_timeslot2.grid(row=9,column=0,columnspan=2,pady=5)
    btn_rem_timeslot2 = Button(f,text='Remove Timeslot',command=lambda: remove_timeslot('sunday'))
    btn_rem_timeslot2.grid(row=9,column=2,columnspan=2,sticky=E)
    timepicker_contents2 = [drop_hour2,lbl_colon2,drop_mints2,drop_period2,btn_add_timeslot2,btn_rem_timeslot2]
    
    e_timeslots2 = ScrolledText(f,width=10,height=6,bg=widgetColor('frame_5'))
    e_timeslots2.grid(row=10,columnspan=6)
    e_timeslots2.configure(state='disable')

    #Number of timeslots entered for week days
    var_numOfTimeslots2 = StringVar()
    var_numOfTimeslots2.set('Entries: 0/'+str(matches_on_sunday))
    lbl_numOfTimeslots2 = Label(f,textvariable=var_numOfTimeslots2,pady=5)
    lbl_numOfTimeslots2.grid(row=11,columnspan=6)

    #Empty Label at end to fix width of parent frame
    lbl_empty2 = Label(f,text=' '*60)
    #lbl_empty2.grid(row=6)
    
    frame_5_contents = [btn_51,lbl_empty,lbl_timeslots1,lbl_numOfTimeslots1,lbl_timeslots2,lbl_numOfTimeslots2,lbl_empty2]
    frame_5_contents += timepicker_contents1
    frame_5_contents += timepicker_contents2
    for child in frame_5_contents:
        try:
            child.configure(state='disable')
        except Exception as e:
            print('frame 5 Error for',type(child))
            print(e)

    colorize_widget('frame_5')
    
    f.grid(row=0,column=3,rowspan=2,padx = 10, pady = 10,sticky=W+N)


def load_leftButton_a():
    global btn_left
    global editing_teamsAndVenues

    editing_teamsAndVenues = True

    try:
        print('Destroying btn_left_a')
        btn_left.destroy()
    except Exception as e:
        print(e)
        print()
    
    btn_left = Button(w,text = 'Lock Teams and Venues',command=lock_teamsAndVenues,cursor='hand2',padx=50,bg=widgetColor('btn_left'))
    btn_left.pack()
    btn_left.place(x=240,y=545,anchor=S)


def load_leftButton_b():
    global btn_left
    global editing_teamsAndVenues

    editing_teamsAndVenues = False

    try:
        print('Destroying btn_left_b')
        btn_left.destroy()
    except Exception as e:
        print(e)
        print()
    
    btn_left = Button(w,text = 'Edit Teams and Venues',command=edit_teamsAndVenues,cursor='hand2',padx=52,bg=widgetColor('btn_left'))
    btn_left.pack()
    btn_left.place(x=240,y=545,anchor=S)


def load_status():
    global lbl_status
    
    lbl_status = Label(w,text='waiting...',bg=widgetColor('status'))
    lbl_status.pack()
    lbl_status.place(x=int(window_x/2),y=560,anchor=N)


def load_scheduleButton():
    global btn_schedule
    import tkinter.font as tkFont
    myfont = tkFont.Font(family='Arial',size=14)
    
    btn_schedule = Button(w,text = 'CREATE SCHEDULE',font=myfont,command=executeSchedule,cursor='hand2',padx=52,bg=widgetColor('btn_schedule'))
    btn_schedule.pack()
    btn_schedule.place(x=896,y=600,anchor=S)
    



























    
#********************************************************   
#******************DECLARE VARIABLES*********************
    
num_of_teams = None
num_of_groups = None

grouplen = None
groups = None
all_teams = None

min_break = None
try_limit = None
rounds = None
first_match_constraint = None
include_dates = None
include_timeslots = None
include_venues = None
include_homegrounds = None
parallel_groups = None

date = None
matches_per_day = None
matches_on_sunday = None

match_duration = None
timeslots_weekday = None
timeslots_sunday = None

venues = None

pending_from_last_round = None

grouped_matches_set = None

all_matches = None





























#******************************************************
#******************MAIN STATEMENTS*********************

#Create main window
w = Tk()
w.title('Tournament Scheduler: Round Robin')
w.configure(background='GREY')
window_x = 1050
window_y = 600
w.geometry(str(window_x) +'x'+ str(window_y))
#w.resizable(False,False)

#Icon
w.iconbitmap("tournament_cup.ico")

#current_frame refers to the one main parent frame currently active in the window 'w'
current_frame = None


#Load main objects into the window
load_frame_mainInput()
load_status()
load_leftButton_a()
load_scheduleButton()


w.after(1,update)


w.mainloop()




